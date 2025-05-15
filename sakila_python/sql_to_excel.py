import os

import sqlalchemy as sa
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy import orm
from dotenv import load_dotenv
from sqlalchemy.ext.automap import automap_base
import pandas as pd
from openpyxl import Workbook

load_dotenv()

url_object = sa.URL.create(drivername='mysql+pymysql',
                           username=os.getenv('DB_USER'),
                           password=os.getenv('DB_PASSWORD'),
                           host=os.getenv('DB_HOST'),
                           port=os.getenv('DB_PORT'),
                           database=os.getenv('DB_NAME')

                           )

engine = sa.create_engine(url_object)

Session = orm.sessionmaker(engine)

Base = automap_base()

Base.prepare(autoload_with=engine,
             classname_for_table=lambda cls, table_name, table_obj: table_name.title().replace('_', ''))

Film = Base.classes.Film
Payment = Base.classes.Payment
Rental = Base.classes.Rental
Inventory = Base.classes.Inventory

with Session() as session:
    query = (
        sa.select(
            Film.film_id,
            Film.title.label("Film title"),
            Film.description.label("Film rating"),
            Film.rating.label("Film rating"),
            Film.rental_rate.label("Film rental rate"),
            Rental.rental_date,
            Payment.payment_date,
            Payment.amount.label("Payment amount")

        ).join(Inventory, Inventory.film_id == Film.film_id)
        .join(Rental, Rental.inventory_id == Inventory.inventory_id)
        .join(Payment, Payment.rental_id == Rental.rental_id)
        .limit(500)
    )

    result_proxy = session.execute(query)  # dane opakowne w obiekt mają do dadkowe funkcje
    results = list(result_proxy.mappings())
    df = pd.DataFrame(results)

wb = Workbook()
wb.remove(wb.active)
chunk_size = 100
header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
header_font = Font(bold=True)

for i in range(0, len(df), chunk_size):
    ws = wb.create_sheet(title=f"Sheet{i // chunk_size} + 1")
    chunk = df.iloc[i:i + chunk_size]
    for row_idx, row in enumerate(dataframe_to_rows(chunk, index=False, header=True), start=1):
        ws.append(row)
        if row_idx == 1:
            for col_idx in range(1, len(row) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font

wb.save("report.xlsx")




